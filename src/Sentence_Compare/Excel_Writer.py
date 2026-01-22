import os
import time
from turtle import title
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter


class Excel_Writer(object):
    ''' Excel_Writer class creates the workbooks and sheets within them.'''
    
    def __init__(self, workbook_path):
        ''' Save the file path (including target file name) and create a 
        workbook to populate with sheets.
        '''
        
        self.workbook_path = workbook_path
        # Create a new workbook
        self.wb = Workbook()
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                    fill_type="solid")
        self.red_font = Font(color="9C0006")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                                    fill_type="solid")
        self.grey_font = Font(color="808080")
        


    def create_original_sheet(self, data, header):
        ''' The workbook comes with one active sheet, so use it.'''
        
        org_ws = self.wb.active  
        org_ws.title = "Original"
        start_row = 1
        
        # Insert multiple rows with multiple columns
        if header is not None:
            org_ws.append(header)
        for row in data:            
            org_ws.append(row)

        if header is not None:
            # Format the first row only
            first_row = org_ws[1]
            # First column should be bold 
            for cell in first_row:
                cell.font = Font(bold=True)  # Set text to bold
                

    def add_sheet(self, data, title, header, cell_width=32, slant_header=True, titles=False):
        ''' The main workhorse, creates a new sheet then takes a lists of rows 
        and adds each row to the sheet. It also adds formatting options based on
        the header, and the parameters passed in.
        '''
        
        new_ws = self.wb.create_sheet(title=title)
        header_rows = 0
        title_cols = 0
        # Insert multiple rows with multiple columns
        if header is not None:
            # if we have row titles move the header over to account for them
            if isinstance(header[0], list):
                for row in header:
                    if titles:
                        new_row = ['']
                        new_row.extend(row)
                        row = new_row 
                    new_ws.append(row)
            else:
                if titles:
                    new_header = ['']
                    new_header.extend(header)
                    header = new_header                 
                new_ws.append(header)
                
        for row in data:
            new_ws.append(row)
                    
        if header is not None:
            if isinstance(header[0], list):
                header_rows = len(header)
                for i in range(len(header)):
                    # Format the header
                    row = new_ws[i + 1]
                    # First column should be bold and optionally slanted
                    for cell in row:
                        cell.font = Font(bold=True)  # Set text to bold
                        if i == 0 and slant_header:
                            cell.alignment = Alignment(textRotation=75, horizontal='center', vertical='center')  # Set rotation angle and alignment
                        elif i == 1:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                header_rows = 1
                row = new_ws[1]
                # First column should be bold and optionally slanted
                for cell in row:
                    cell.font = Font(bold=True)  # Set text to bold
                    if slant_header:
                        cell.alignment = Alignment(textRotation=75, horizontal='center', vertical='center')  # Set rotation angle and alignment
                
        # Nasty Hack but it makes it more readible.
        if title == 'Summary':  
            self.format_summary_sheet(new_ws)
            
        if titles:
            title_cols = 1
            # First col should be bold
            for row in new_ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)  # Set text to bold
                    break

        # Auto size columns
        for col in new_ws.columns:
            cell_strs = [str(cell.value) for cell in col]
            max_len = 0
            max_len = len(max(cell_strs, key=len))
            adjusted_width = (max_len + 2) 
            if cell_width != 0:
                adjusted_width = min(adjusted_width, cell_width)
            new_ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width


    def add_aldos_sheet(self, data, title, cell_width=50, header=None, slant_header=False, titles=False):
        ''' The sheets Aldo wanted required little different formatting.''' 
        
        new_ws = self.wb.create_sheet(title=title)
        cells2paint = [2, 4, 5]
        # Insert multiple rows with multiple columns
        if header is not None:
            for row in header:
                new_ws.append(row)
        for row in data:    
            source, topic, vol, topic_vol, cos, lu_row, sort = row
            new_ws.append([source, topic, vol, topic_vol, cos, lu_row, sort])

        if header: 
            # Format the first row only
            first_row = new_ws[1]
            # First column should be bold and slanted
            for cell in first_row:
                cell.font = Font(bold=True)  # Set text to bold

        # Auto size columns
        for col in new_ws.columns:
            cell_strs = [str(cell.value) for cell in col]
            max_len = 0
            max_len = len(max(cell_strs, key=len))
            adjusted_width = (max_len + 2) 
            if cell_width != 0:
                adjusted_width = min(cell_width, adjusted_width)
            new_ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
            

    def format_summary_sheet(self, ws):
        ''' Format for the summary information in the last 6 rows.'''
        
        last_row = ws.max_row
        for cell in ws[last_row-5]:
            if cell.value and 'Threshold used:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row-5, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
        for cell in ws[last_row - 4]:
            if cell.value and 'Consolidation level:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row-4, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
        for cell in ws[last_row - 3]:
            if cell.value and 'Total Records:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row-3, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
        for cell in ws[last_row - 2]:
            if cell.value and 'Initial Topics:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row-2, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
        for cell in ws[last_row - 1]:
            if cell.value and 'Consolidated Topics:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row-1, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
        for cell in ws[last_row]:
            if cell.value and 'Consolidation factor:' in cell.value:
                cell.font = Font(bold=True)
                cell_to_right = ws.cell(row=last_row, column=cell.column + 1)
                cell_to_right.font = Font(bold=True)
                break
 

    def format_values_below_threshold(self, ws, data_start, threshold):
        ''' Apply conditional formatting to cosine tables, grey zeros and red 
        below values. 
        '''
        
        ''' Must happen after the summary sheet is created so the cell 
        containing the threshold value can be addressed.
        '''
        
        threshold_loc = self.locate_threshold()
        first_col, first_row = data_start        
        first_col_letter = get_column_letter(first_col)
        last_row = ws.max_row
        last_col_letter = get_column_letter(ws.max_column)
        area_to_apply = format(f'${first_col_letter}${first_row}:${last_col_letter}${last_row}')
        
        # Create a conditional formatting rule
        # rule = CellIsRule(operator='lessThan', formula=[str(threshold)], fill=self.red_fill, font=self.red_font)
        threshold_rule = CellIsRule(operator='lessThan', formula=[threshold_loc], fill=self.red_fill, font=self.red_font)
        zero_rule = CellIsRule(operator='equal', formula=[0], fill=self.white_fill, font=self.grey_font)

        # If not applied in this order the zero_rule won't take effect.
        ws.conditional_formatting.add(area_to_apply, zero_rule)
        ws.conditional_formatting.add(area_to_apply, threshold_rule)


    def format_tracking_below_threshold(self, ws_tracking, trck_start, values_title, threshold):
        ''' Apply conditional formatting to tracking tables using indirect 
        addressing to get the cosine values for an index.
        '''
        
        ''' Must happen after the summary sheet is created so the cell 
        containing the threshold value can be addressed.
        '''

        threshold_loc = self.locate_threshold()
        first_col, first_row = trck_start        
        first_col_letter = get_column_letter(first_col)
        last_row = ws_tracking.max_row
        last_col_letter = get_column_letter(ws_tracking.max_column)
        area_to_apply = format(f'${first_col_letter}${first_row}:${last_col_letter}${last_row}')

        
        formula = f'INDIRECT(\"{values_title}!\"&ADDRESS(ROW(),COLUMN())) < {threshold_loc}'

        # Apply the conditional formatting rule to the specified range
        threshold_rule = FormulaRule(formula=[formula], fill=self.red_fill, font=self.red_font)
        zero_rule = CellIsRule(operator='equal', formula=[0], fill=self.white_fill, font=self.grey_font)
        
        # If not applied in this order the zero_rule won't take effect.
        ws_tracking.conditional_formatting.add(area_to_apply, zero_rule)
        ws_tracking.conditional_formatting.add(area_to_apply, threshold_rule)


    def format_aldos_conditionals(self):
        ''' Add conditional formatting to Aldo's sheet. '''
        
        ws = self.wb['Aldos']
        fst_row = 2
        lst_row = ws.max_row
        threshold_loc = self.locate_threshold()
        
        ws.conditional_formatting.add(f'$E${fst_row}:$E${lst_row}',
            CellIsRule(operator='lessThan', formula=[threshold_loc], 
            stopIfTrue=False, font=self.red_font, fill=self.red_fill))
        cells = format(f'$B${fst_row}:$B${lst_row} $D${fst_row}:$D${lst_row} $F${fst_row}:$G${lst_row}')
        ws.conditional_formatting.add(cells,
            FormulaRule(formula=[f'INDIRECT(ADDRESS(ROW(), 5)) < {threshold_loc}'],
            stopIfTrue=False, font=self.red_font, fill=self.red_fill))



    def apply_conditional_formatting(self):
        ''' Which conditional formatting to apply to the worksheets.'''
        self.format_aldos_conditionals()
        
        threshold_loc = self.locate_threshold()
        sheet_titles = self.wb.sheetnames
        
        cos_sheets_info = {}
        cos_sheets_info['Sorted'] = ((1, 2))
        cos_sheets_info['Best'] = ((1, 3))
        cos_sheets_info['Consolidated'] = ((1, 3))
        
        for cos_name, loc in cos_sheets_info.items():
            if cos_name in sheet_titles:
                ws = self.wb[cos_name]                
                self.format_values_below_threshold(ws, loc, threshold_loc)
 
        trck_sheets_info = {}        
        trck_sheets_info['Sorted_Lookup'] = ('Sorted', (1, 2))
        trck_sheets_info['Best_Tracking'] = ('Best', (1, 3))
        trck_sheets_info['Consolidated_Tracking'] = ('Consolidated', (1, 3))
        trck_sheets_info['Sentences'] = ('Consolidated', (1, 3))

        for trck_name, items in trck_sheets_info.items():
            if trck_name in sheet_titles:
                cos_name, start_loc = items
                ws = self.wb[trck_name]
                self.format_tracking_below_threshold(ws, start_loc,
                                                     cos_name, threshold_loc)

    def locate_threshold(self):
        ''' Find the threshold value in the Summary sheet.  '''
        
        ''' This location will change based on the the number of CSV files read.'''
        
        ws = self.wb['Summary']
        threshold_row = ws.max_row - 5
        return f'Summary!$C${threshold_row}'


    def write_workbook(self):
        ''' Apply final formatting and save the workbook.'''
        
        # All the sheets should be added so dependancies between sheets should 
        # not cause a problem.
        self.apply_conditional_formatting()
        # Create the path if needed
        os.makedirs(os.path.dirname(self.workbook_path), exist_ok=True)
        # Save the workbook
        self.wb.save(self.workbook_path)



 



